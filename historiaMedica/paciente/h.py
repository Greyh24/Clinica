import tkinter as tk
from tkinter import filedialog
from tkcalendar import DateEntry
from tkinter import ttk
from time import strftime
from PIL import Image, ImageTk

class TuVentana:
    def __init__(self, root):
        self.root = root
        self.scrollable_frame = tk.Frame(root)
        self.scrollable_frame.pack(expand=True, fill='both')

        # Crear un frame para datos actuales
        datos_frame = tk.Frame(self.scrollable_frame, bg='#8C9BBA')
        datos_frame.grid(column=0, row=0, padx=5, pady=5, sticky='ew')

        
        self.lblDM = tk.Label(datos_frame, text='Diagnóstico Médico: ', font=('ARIAL', 24, 'bold'),fg='#ffffff', bg='#8C9BBA')
        self.lblDM.grid(column=2, row=0, columnspan=3, padx=5, pady=5)

        self.lblPaci = tk.Label(datos_frame, text='Paciente: ', font=('ARIAL', 14, 'bold'),fg='#ffffff', bg='#8C9BBA')
        self.lblPaci.grid(column=0, row=1, padx=0, pady=5)

    
        # Crear un frame para la tabla
        tabla_frame = tk.Frame(self.scrollable_frame, bg='#8C9BBA')
        tabla_frame.grid(column=0, row=1, padx=5, pady=5, sticky='ew')

            # Crear una tabla
        columns = ('Columna 1', 'Columna 2', 'Columna 3')
        self.tabla = ttk.Treeview(tabla_frame, columns=columns, show='headings')

        for col in columns:
            self.tabla.heading(col, text=col)

        self.tabla.grid(column=0, row=0, sticky='nsew')

        # Configurar el scrollbar para la tabla
        scrollbar = ttk.Scrollbar(tabla_frame, orient='vertical', command=self.tabla.yview)
        scrollbar.grid(column=1, row=0, sticky='ns')
        self.tabla.configure(yscrollcommand=scrollbar.set)

            # Agregar datos de ejemplo a la tabla
        data = [('Dato 1', 'Dato 2', 'Dato 3'),
                    ('Dato 4', 'Dato 5', 'Dato 6'),
                    ('Dato 7', 'Dato 8', 'Dato 9')]

        for row in data:
            self.tabla.insert('', 'end', values=row)

    def on_canvas_configure(self, event):
        # Configurar la región de desplazamiento del canvas cuando cambia el tamaño
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def on_frame_configure(self, event):
        # Configurar la región de desplazamiento del frame interior cuando cambia el tamaño
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def validate_number(self, value):
        # Función de validación para aceptar solo números
        return value.isdigit() or value == ""
    
    def validate_decimal(self, value):
        # Función de validación para aceptar números decimales
        try:
            float(value)
            return True
        except ValueError:
            return False
    
    def validate_letter(self, value):
        # Función de validación para aceptar solo letras
        return all(c.isalpha() or c.isspace() for c in value) or value == ""
    
    def actualizar_hora(self):
        hora_actual = strftime('%H:%M %p')
        self.entryHoraP.delete(0, tk.END)
        self.entryHoraP.insert(0, hora_actual)

    def actualizar_hora2(self):
        hora_actual = strftime('%H:%M %p')
        self.entryHoraAT.delete(0, tk.END)
        self.entryHoraAT.insert(0, hora_actual)  

    def actualizar_hora3(self):
        hora_actual = strftime('%H:%M %p')
        self.entryHoraD.delete(0, tk.END)
        self.entryHoraD.insert(0, hora_actual)   

    def mostrar_imagen(self, event, contenedor):
        # Abrir el cuadro de diálogo de selección de archivo
        ruta_imagen = filedialog.askopenfilename(title="Seleccionar Imagen", filetypes=[("Archivos de imagen", "*.png;*.jpg;*.jpeg")])

        # Verificar si se seleccionó una imagen
        if ruta_imagen:
            imagen = Image.open(ruta_imagen)
            self.mostrar_imagen_en_label(imagen, contenedor)

    def mostrar_imagen_en_label(self, imagen, contenedor):
        imagen.thumbnail((150, 100))
        imagen_tk = ImageTk.PhotoImage(imagen)

        contenedor.config(image=imagen_tk)
        contenedor.image = imagen_tk  # Mantener una referencia para evitar que sea eliminado

        if contenedor == self.contenedor_MedRes:
            self.imagen_tk_MedRes = imagen_tk
        elif contenedor == self.contenedor_ExaAdi:
            self.imagen_tk_ExaAdi = imagen_tk

    def mostrar_cuadro_dialogo(self, event):
        # Abrir el cuadro de diálogo de selección de archivo
        ruta_imagen = filedialog.askopenfilename(title="Seleccionar Imagen", filetypes=[("Archivos de imagen", "*.png;*.jpg;*.jpeg")])

        # Verificar si se seleccionó una imagen
        if ruta_imagen:
            # Agregar la referencia de la imagen a la lista
            self.lista_imagenes.insert(tk.END, ruta_imagen)

            # Seleccionar la última imagen agregada
            self.lista_imagenes.selection_clear(0, tk.END)
            self.lista_imagenes.selection_set(tk.END)

            # Mostrar la última imagen en el contenedor
            self.seleccionar_imagen()

    def seleccionar_imagen(self, event=None):
        # Obtener el índice de la imagen seleccionada en el Listbox
        seleccion = self.lista_imagenes.curselection()

        if seleccion:
            indice = int(seleccion[0])
            ruta_imagen = self.lista_imagenes.get(indice)

            # Abrir la imagen con Pillow
            imagen = Image.open(ruta_imagen)

            self.mostrar_imagen_en_label(imagen, self.contenedor_ExaAdi)

if __name__ == "__main__":
    root = tk.Tk()
    app = TuVentana(root)
    root.mainloop()


import tkinter as tk
from tkinter import filedialog, messagebox
from tkcalendar import DateEntry
from tkinter import ttk
import openpyxl
from datetime import datetime
import csv

class VentanaAdultos(tk.Frame):
    def __init__(self, root):
        super().__init__(root, width=1250, height=620)
        self.root = root

        self.frame_info = tk.Frame(self, bg='#CDD8FF')
        self.frame_info.grid(row=0, column=0, sticky="nsew")

        self.frame_tabla = tk.Frame(self, bg='#CDD8FF')
        self.frame_tabla.grid(row=1, column=0, sticky="nsew")

        self.frame_tabla_visualizacion = tk.Frame(self, bg='#CDD8FF')
        self.frame_tabla_visualizacion.grid(row=2, column=0, sticky="nsew")

        self.datos_originales = []

        self.inicializar_gui()

    def inicializar_gui(self):
        self.campos_paciente()
        self.crear_tabla()
        self.cargar_formulario()

    def campos_paciente(self):
        self.scrollbar_y = tk.Scrollbar(self.frame_info, orient="vertical", command=self.yview)
        self.scrollbar_y.grid(row=0, column=1, sticky="ns")

        self.canvas = tk.Canvas(self.frame_info, bg='#CDD8FF', yscrollcommand=self.scrollbar_y.set)
        self.canvas.grid(row=0, column=0, sticky="nsew")

        self.scrollable_frame = tk.Frame(self.canvas, bg='#CDD8FF')
        self.scrollable_frame.grid(row=0, column=0, sticky="nsew")

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor=tk.NW)

        self.canvas.bind("<Configure>", self.on_canvas_configure)
        self.scrollable_frame.bind("<Configure>", self.on_frame_configure)

        self.frame_info.columnconfigure(0, weight=2, minsize=900)

        self.lblHC = tk.Label(self.scrollable_frame, text='Historia Clinica: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblHC.grid(column=0, row=0, padx=5, pady=5)

        self.lblfecha = tk.Label(self.scrollable_frame, text='Fecha: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblfecha.grid(column=2, row=0, padx=5, pady=5)

        self.lbledad = tk.Label(self.scrollable_frame, text='Edad: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lbledad.grid(column=0, row=1, padx=5, pady=5)

        self.lblsexo = tk.Label(self.scrollable_frame, text='Sexo: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblsexo.grid(column=2, row=1, padx=5, pady=5)

        self.lblNombre = tk.Label(self.scrollable_frame, text='Nombre: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblNombre.grid(column=0, row=2, padx=5, pady=5)

        self.lblApellido = tk.Label(self.scrollable_frame, text='Apellido: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblApellido.grid(column=2, row=2, padx=5, pady=5)

        self.lblFechaN = tk.Label(self.scrollable_frame, text='Fecha de Nacimiento: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblFechaN.grid(column=0, row=3, padx=5, pady=5)

        self.lblDNI = tk.Label(self.scrollable_frame, text='DNI: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblDNI.grid(column=2, row=3, padx=5, pady=5)

        self.lbltelef = tk.Label(self.scrollable_frame, text='Telefono: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lbltelef.grid(column=0, row=4, padx=5, pady=5)

        self.lblDNI = tk.Label(self.scrollable_frame, text='Buscar por Hc: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblDNI.grid(column=0, row=6, padx=5, pady=5)

        self.lbltelef = tk.Label(self.scrollable_frame, text='Buscar por DNI: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lbltelef.grid(column=3, row=6, padx=5, pady=15)

        self.svHC = tk.StringVar()
        self.entryHC = tk.Entry(self.scrollable_frame, textvariable=self.svHC, width=20, font=('ARIAL', 10))
        self.entryHC.grid(column=1, row=0, padx=5, pady=5)

        self.svfecha = tk.StringVar()
        self.entryfecha = DateEntry(self.scrollable_frame, textvariable=self.svfecha, width=20, font=('ARIAL', 10), date_pattern='dd-mm-yyyy')
        self.entryfecha.grid(column=3, row=0, padx=5, pady=5)

        self.svedad = tk.StringVar()
        self.entryedad = tk.Entry(self.scrollable_frame, textvariable=self.svedad, width=20, font=('ARIAL', 10))
        self.entryedad.grid(column=1, row=1, padx=5, pady=5)

        self.svsexo_m = tk.IntVar()
        self.checkbox_masculino = tk.Checkbutton(self.scrollable_frame, text='H', variable=self.svsexo_m,font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.checkbox_masculino.grid(column=3, row=1, padx=(0, 1), pady=5)

        self.svsexo_f = tk.IntVar()
        self.checkbox_femenino = tk.Checkbutton(self.scrollable_frame, text='M', variable=self.svsexo_f,font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.checkbox_femenino.grid(column=3, row=1, padx=(90, 0), pady=5)

        self.svNombre = tk.StringVar()
        self.entryNombre = tk.Entry(self.scrollable_frame, textvariable=self.svNombre, width=20, font=('ARIAL', 10))
        self.entryNombre.grid(column=1, row=2, padx=5, pady=5)

        self.svApellido = tk.StringVar()
        self.entryApellido = tk.Entry(self.scrollable_frame, textvariable=self.svApellido, width=20, font=('ARIAL', 10))
        self.entryApellido.grid(column=3, row=2, padx=5, pady=5)

        self.svFechaN = tk.StringVar()
        self.entryFechaN = DateEntry(self.scrollable_frame, textvariable=self.svFechaN, width=20, font=('ARIAL', 10), date_pattern='dd-mm-yyyy')
        self.entryFechaN.grid(column=1, row=3, padx=5, pady=5)

        self.svDNI = tk.StringVar()
        self.entryDNI = tk.Entry(self.scrollable_frame, textvariable=self.svDNI, width=20, font=('ARIAL', 10))
        self.entryDNI.grid(column=3, row=3, padx=5, pady=5)

        self.svtelef = tk.StringVar()
        self.entrytelef = tk.Entry(self.scrollable_frame, textvariable=self.svtelef, width=20, font=('ARIAL', 10))
        self.entrytelef.grid(column=1, row=4, padx=5, pady=5)

        self.svBuscHC = tk.StringVar()
        self.entryBuscHC = tk.Entry(self.scrollable_frame, textvariable=self.svBuscHC, width=20, font=('ARIAL', 10))
        self.entryBuscHC.grid(column=1, row=6, padx=5, pady=5)

        self.svBuscDNI = tk.StringVar()
        self.entryBuscDNI = tk.Entry(self.scrollable_frame, textvariable=self.svBuscDNI, width=20, font=('ARIAL', 10))
        self.entryBuscDNI.grid(column=4, row=6, padx=5, pady=15)

        self.btnGuardar = tk.Button(self.scrollable_frame, text='Guardar', width=10, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#158645', cursor='hand2', activebackground='#35BD6F',command=self.guardar_datos)
        self.btnGuardar.grid(column=1, row=5, padx=5, pady=5)

        self.btnModificar = tk.Button(self.scrollable_frame, text='Modificar', width=10, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#1658A2', cursor='hand2', activebackground='#3D69F0',command=self.modificar_seleccion)
        self.btnModificar.grid(column=2, row=5, padx=5, pady=5)

        self.btnEliminar = tk.Button(self.scrollable_frame, text='Eliminar', width=10, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#D32F2F', cursor='hand2', activebackground='#E72D40',command=self.eliminar_seleccion)
        self.btnEliminar.grid(column=3, row=5, padx=5, pady=5)

        self.btnImportar = tk.Button(self.scrollable_frame, text='Importar', width=10, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#CF811E', cursor='hand2', activebackground='#E72D40',command=self.importar_excel)
        self.btnImportar.grid(column=4, row=5, padx=5, pady=5)

        self.btnExportar = tk.Button(self.scrollable_frame, text='Exportar', width=10, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#CF811E', cursor='hand2', activebackground='#E72D40',command=self.exportar_a_excel)
        self.btnExportar.grid(column=5, row=5, padx=15, pady=5)

        self.btnBuscarhc = tk.Button(self.scrollable_frame, text='Buscar', width=10, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#0F1010', cursor='hand2', activebackground='#FFFEFE',command=self.filtrar_tabla)
        self.btnBuscarhc.grid(column=2, row=6, padx=15, pady=15)

        self.btnBuscardni = tk.Button(self.scrollable_frame, text='Buscar', width=10, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#0F1010', cursor='hand2', activebackground='#FFFEFE',command=self.filtrar_tabla)
        self.btnBuscardni.grid(column=5, row=6, padx=5, pady=15)

    def crear_tabla(self):
        # Crear la tabla con el widget Treeview
        self.tabla = ttk.Treeview(self.frame_tabla, columns=("HC", "Fecha", "Edad", "Sexo", "Nombre", "Apellido", "FechaN", "DNI", "Telefono"))

        # Definir encabezados
        self.tabla.heading("#0", text="ID")
        self.tabla.heading("HC", text="Historia Clínica")
        self.tabla.heading("Fecha", text="Fecha")
        self.tabla.heading("Edad", text="Edad")
        self.tabla.heading("Sexo", text="Sexo")
        self.tabla.heading("Nombre", text="Nombre")
        self.tabla.heading("Apellido", text="Apellido")
        self.tabla.heading("FechaN", text="Fecha de Nacimiento")
        self.tabla.heading("DNI", text="DNI")
        self.tabla.heading("Telefono", text="Teléfono")

        # Configurar las columnas
        self.tabla.column("#0", stretch=tk.NO, width=0)  # ID
        self.tabla.column("HC", anchor=tk.CENTER, width=100)
        self.tabla.column("Fecha", anchor=tk.CENTER, width=100)
        self.tabla.column("Edad", anchor=tk.CENTER, width=50)
        self.tabla.column("Sexo", anchor=tk.CENTER, width=50)
        self.tabla.column("Nombre", anchor=tk.W, width=100)
        self.tabla.column("Apellido", anchor=tk.W, width=100)
        self.tabla.column("FechaN", anchor=tk.CENTER, width=100)
        self.tabla.column("DNI", anchor=tk.CENTER, width=80)
        self.tabla.column("Telefono", anchor=tk.CENTER, width=80)

        # Configurar la barra de desplazamiento
        scrollbar_y = ttk.Scrollbar(self.frame_tabla, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscroll=scrollbar_y.set)
        scrollbar_y.grid(row=0, column=1, sticky="ns")

        # Ubicar la tabla en el grid
        self.tabla.grid(row=0, column=0, sticky="nsew")

        # Configurar la barra de desplazamiento horizontal
        self.frame_tabla.columnconfigure(0, weight=1)
        self.frame_tabla.rowconfigure(0, weight=1)

        # Manejar la selección de fila
        self.tabla.bind("<ButtonRelease-1>", self.on_select)

    def cargar_desde_excel(self):
        try:
            archivo_excel = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")])
            if archivo_excel:
                workbook = openpyxl.load_workbook(archivo_excel)
                sheet = workbook.active

                self.limpiar_campos_busqueda()

                self.datos_originales = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    paciente = {
                        'HC': row[0],
                        'Fecha': row[1],
                        'Edad': row[2],
                        'Sexo': row[3],
                        'Nombre': row[4],
                        'Apellido': row[5],
                        'FechaN': row[6],
                        'DNI': row[7],
                        'Telefono': row[8],
                    }
                    self.datos_originales.append(paciente)

                self.restablecer_tabla()
                messagebox.showinfo("Éxito", "Datos cargados desde el archivo Excel correctamente.")

        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar datos desde el archivo Excel: {str(e)}")

    def actualizar_tabla(self):
        # Clear the table before updating
        self.tabla.delete(*self.tabla.get_children())

        # Update the table with the latest data
        for paciente in self.datos_originales:
            row_data = [
                paciente['HC'], paciente['Fecha'], paciente['Edad'],
                paciente['Sexo'], paciente['Nombre'], paciente['Apellido'],
                paciente['FechaN'], paciente['DNI'], paciente['Telefono']
            ]
            self.tabla.insert("", "end", values=row_data, tags=(paciente['HC'],))

    def guardar_datos(self):
        try:
            hc = self.svHC.get()
            fecha = self.svfecha.get()
            edad = self.svedad.get()
            sexo = 'H' if self.svsexo_m.get() else 'M'
            nombre = self.svNombre.get()
            apellido = self.svApellido.get()
            fecha_nacimiento = self.svFechaN.get()
            dni = self.svDNI.get()
            telefono = self.svtelef.get()

            # Validar campos antes de guardar
            if not hc or not fecha or not edad or not sexo or not nombre or not apellido or not fecha_nacimiento or not dni or not telefono:
                messagebox.showwarning("Advertencia", "Todos los campos son obligatorios.")
                return

            nuevo_paciente = {
                'HC': hc,
                'Fecha': fecha,
                'Edad': edad,
                'Sexo': sexo,
                'Nombre': nombre,
                'Apellido': apellido,
                'FechaN': fecha_nacimiento,
                'DNI': dni,
                'Telefono': telefono
            }

            # Verificar si ya existe un paciente con la misma Historia Clínica
            if any(paciente['HC'] == hc for paciente in self.datos_originales):
                messagebox.showwarning("Advertencia", "Ya existe un paciente con la misma Historia Clínica.")
                return

            # Guardar en la lista de datos_originales
            self.datos_originales.append(nuevo_paciente)

            # Insertar en la tabla y guardar en el archivo CSV
            row_data = [nuevo_paciente[field] for field in ['HC', 'Fecha', 'Edad', 'Sexo', 'Nombre', 'Apellido', 'FechaN', 'DNI', 'Telefono']]
            self.tabla.insert("", "end", values=row_data, tags=(hc,))
            self.guardar_en_csv(row_data)

            # Limpiar los campos después de guardar
            for entry in [self.entryHC, self.entryfecha, self.entryedad,
                        self.entryNombre, self.entryApellido, self.entryFechaN, self.entryDNI, self.entrytelef]:
                entry.delete(0, tk.END)

            # Uncheck the checkboxes after saving
            self.svsexo_m.set(0)
            self.svsexo_f.set(0)

        except Exception as e:
            print("Error", f"Error al guardar datos: {str(e)}")

    def guardar_en_csv(self, datos):
        try:
            with open("datos_adultos.csv", mode="a", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(datos)
        except Exception as e:
            print("Error", f"Error al guardar en CSV: {str(e)}")

    def cargar_formulario(self):
        try:
            with open("datos_adultos.csv", mode="r", encoding="utf-8") as file:
                reader = csv.reader(file)
                next(reader, None)  # Skip header
                datos = [row for row in reader]
                self.datos_originales = []

                # Limpiar la tabla antes de cargar los datos
                self.tabla.delete(*self.tabla.get_children())

                for row in datos:
                    paciente = {
                        'HC': row[0],
                        'Fecha': row[1],
                        'Edad': row[2],
                        'Sexo': row[3],
                        'Nombre': row[4],
                        'Apellido': row[5],
                        'FechaN': row[6],
                        'DNI': row[7],
                        'Telefono': row[8]
                    }
                    self.datos_originales.append(paciente)

                    # Insertar en la tabla
                    self.tabla.insert("", "end", values=row, tags=(row[0],))

        except FileNotFoundError:
            pass
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar formulario: {str(e)}")

    def eliminar_seleccion(self):
            item_seleccionado = self.tabla.selection()
            if item_seleccionado:
                hc_seleccionado = self.tabla.item(item_seleccionado)['values'][0]
                self.tabla.tag_configure(hc_seleccionado, background="")
                self.tabla.delete(item_seleccionado)
                
    def modificar_seleccion(self):
       pass

    def filtrar_tabla(self):
        try:
            hc_buscar = self.entryBuscHC.get()
            dni_buscar = self.entryBuscDNI.get()

            self.tabla.delete(*self.tabla.get_children())

            for datos_fila in self.datos_originales:
                hc_fila = datos_fila[0]
                dni_fila = datos_fila[8]

                if ((not hc_buscar or hc_fila == hc_buscar) and 
                    (not dni_buscar or dni_fila == dni_buscar)):
                    self.tabla.insert("", "end", values=datos_fila, tags=(datos_fila[0],))
        except Exception as e:
            messagebox.showerror("Error", f"Error al filtrar tabla: {str(e)}")

    def limpiar_campos_busqueda(self):
        try:
            self.entryBuscHC.delete(0, tk.END)
            self.entryBuscDNI.delete(0, tk.END)
            self.filtrar_tabla()
        except Exception as e:
            messagebox.showerror("Error", f"Error al limpiar campos de búsqueda: {str(e)}")

    def restablecer_tabla(self):
        try:
            self.tabla.selection_remove(self.tabla.selection())
            for i, fila_tabla in enumerate(self.tabla.get_children()):
                self.tabla.reattach(fila_tabla, "", i)
        except Exception as e:
            messagebox.showerror("Error", f"Error al restablecer la tabla: {str(e)}")

    def importar_excel(self):
        try:
            archivo_excel = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx"), ("Archivos CSV", "*.csv")])

            if archivo_excel:
                workbook = openpyxl.load_workbook(archivo_excel)
                sheet = workbook.active

                claves_existente = set()  # Conjunto para almacenar las claves existentes en la tabla

                for fila_tabla in self.tabla.get_children():
                    clave_existente = self.tabla.item(fila_tabla)['values'][0]
                    claves_existente.add(clave_existente)

                datos_temporales = []

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] not in claves_existente and row not in self.datos_originales:
                        claves_existente.add(row[0])
                        datos_temporales.append(row)
                        self.tabla.insert("", "end", values=row, tags=(row[0],))
                        self.guardar_en_csv(row)

                workbook.close()

                # Limpiar la lista datos_temporales antes de importar de nuevo
                datos_temporales.clear()

                # Seleccionar la última fila insertada después de importar
                ultima_fila = self.tabla.get_children()[-1]
                self.tabla.selection_set(ultima_fila)

                # Llamar a la función de selección de fila después de importar
                self.seleccionar_fila_desde_importar()
        except Exception as e:
            messagebox.showerror("Error", f"Error al importar desde Excel: {str(e)}")

    def exportar_a_excel(self):
       pass

    def seleccionar_fila_desde_importar(self):
       pass
    def on_canvas_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
      
    def on_frame_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def yview(self, *args):
        try:
            self.canvas.yview(*args)
        except Exception as e:
            messagebox.print("Error", f"Error en la configuración de la vista vertical: {str(e)}")

    def validate_number(self, value):
        # Función de validación para aceptar solo números
        return value.isdigit() or value == ""

    def validate_letter(self, value):
        # Función de validación para aceptar solo letras
        return all(c.isalpha() or c.isspace() for c in value) or value == ""

    def on_select(self, event):

        # Get the selected item
        selected_item = self.tabla.selection()

        if selected_item:
            # Do something with the selected item
            print(f"Selected item: {self.tabla.item(selected_item)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = VentanaAdultos(root)
    app.mainloop()


