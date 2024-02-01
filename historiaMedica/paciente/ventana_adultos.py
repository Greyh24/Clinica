import tkinter as tk
from tkinter import filedialog, messagebox
from tkcalendar import DateEntry
from tkinter import ttk
import openpyxl
import csv

class VentanaAdultos(tk.Frame):
    def __init__(self, root):
        super().__init__(root, width=1250, height=620)
        self.root = root

        self.frame_info = tk.Frame(self, bg='#CDD8FF')
        self.frame_info.grid(row=0, column=0, sticky="nsew")

        self.frame_tabla = tk.Frame(self, bg='#CDD8FF')
        self.frame_tabla.grid(row=1, column=0, sticky="nsew")

        self.inicializar_gui()
        self.selected_index = None

    def inicializar_gui(self):
        self.campos_paciente()
        self.crear_tabla()

    def on_canvas_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def on_frame_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def yview(self, *args):
        pass

    def validate_number(self, value):
        return value.isdigit() or value == ""

    def validate_letter(self, value):
        return all(c.isalpha() or c.isspace() for c in value) or value == ""
    
    def campos_paciente(self):
        self.scrollbar_y = tk.Scrollbar(self.frame_info, orient="vertical", command=self.yview)
        self.scrollbar_y.grid(row=0, column=1, sticky="ns")

        self.canvas = tk.Canvas(self.frame_info, bg='#CDD8FF', yscrollcommand=self.scrollbar_y.set)
        self.canvas.grid(row=0, column=0, sticky="nsew")

        self.scrollable_frame = tk.Frame(self.canvas, bg='#CDD8FF')
        self.scrollable_frame.grid(row=0, column=0, sticky="nsew")

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor=tk.NW)

        self.canvas.bind("<Configure>", self.on_canvas_configure)
        self.scrollable_frame.bind("<Configure>", self.on_frame_configure)

        self.frame_info.columnconfigure(0, weight=2, minsize=900)

        self.lblHC = tk.Label(self.scrollable_frame, text='Historia Clinica: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblHC.grid(column=0, row=0, padx=5, pady=5)

        self.lblfecha = tk.Label(self.scrollable_frame, text='Fecha: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblfecha.grid(column=2, row=0, padx=5, pady=5)

        self.lbledad = tk.Label(self.scrollable_frame, text='Edad: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lbledad.grid(column=0, row=1, padx=5, pady=5)

        self.lblsexo = tk.Label(self.scrollable_frame, text='Sexo: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblsexo.grid(column=2, row=1, padx=5, pady=5)

        self.lblNombre = tk.Label(self.scrollable_frame, text='Nombre: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblNombre.grid(column=0, row=2, padx=5, pady=5)

        self.lblApellido = tk.Label(self.scrollable_frame, text='Apellido: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblApellido.grid(column=2, row=2, padx=5, pady=5)

        self.lblFechaN = tk.Label(self.scrollable_frame, text='Fecha de Nacimiento: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblFechaN.grid(column=0, row=3, padx=5, pady=5)

        self.lblDNI = tk.Label(self.scrollable_frame, text='DNI: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblDNI.grid(column=2, row=3, padx=5, pady=5)

        self.lbltelef = tk.Label(self.scrollable_frame, text='Telefono: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lbltelef.grid(column=0, row=4, padx=5, pady=5)

        self.lblDNI = tk.Label(self.scrollable_frame, text='Buscar por Hc: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lblDNI.grid(column=0, row=5, padx=5, pady=8)

        self.lbltelef = tk.Label(self.scrollable_frame, text='Buscar por DNI: ', font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.lbltelef.grid(column=0, row=6, padx=5, pady=8)

        self.svHC = tk.StringVar()
        self.entryHC = tk.Entry(self.scrollable_frame, textvariable=self.svHC, width=20, font=('ARIAL', 10))
        self.entryHC.grid(column=1, row=0, padx=5, pady=5)

        self.svfecha = tk.StringVar()
        self.entryfecha = DateEntry(self.scrollable_frame, textvariable=self.svfecha, width=20, font=('ARIAL', 10), date_pattern='dd-mm-yyyy')
        self.entryfecha.grid(column=3, row=0, padx=5, pady=5)

        self.svedad = tk.StringVar()
        self.entryedad = tk.Entry(self.scrollable_frame, textvariable=self.svedad, width=20, font=('ARIAL', 10))
        self.entryedad.grid(column=1, row=1, padx=5, pady=5)

        self.svsexo_m = tk.IntVar()
        self.checkbox_masculino = tk.Checkbutton(self.scrollable_frame, text='M', variable=self.svsexo_m, font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.checkbox_masculino.grid(column=3, row=1, padx=(0, 1), pady=5)

        self.svsexo_f = tk.IntVar()
        self.checkbox_femenino = tk.Checkbutton(self.scrollable_frame, text='F', variable=self.svsexo_f,font=('ARIAL', 10, 'bold'), bg='#CDD8FF')
        self.checkbox_femenino.grid(column=3, row=1, padx=(90, 0), pady=5)

        self.svNombre = tk.StringVar()
        self.entryNombre = tk.Entry(self.scrollable_frame, textvariable=self.svNombre, width=20, font=('ARIAL', 10))
        self.entryNombre.grid(column=1, row=2, padx=5, pady=5)

        self.svApellido = tk.StringVar()
        self.entryApellido = tk.Entry(self.scrollable_frame, textvariable=self.svApellido, width=20, font=('ARIAL', 10))
        self.entryApellido.grid(column=3, row=2, padx=5, pady=5)

        self.svFechaN = tk.StringVar()
        self.entryFechaN = DateEntry(self.scrollable_frame, textvariable=self.svFechaN, width=20, font=('ARIAL', 10), date_pattern='dd-mm-yyyy')
        self.entryFechaN.grid(column=1, row=3, padx=5, pady=5)

        self.svDNI = tk.StringVar()
        self.entryDNI = tk.Entry(self.scrollable_frame, textvariable=self.svDNI, width=20, font=('ARIAL', 10))
        self.entryDNI.grid(column=3, row=3, padx=5, pady=5)

        self.svtelef = tk.StringVar()
        self.entrytelef = tk.Entry(self.scrollable_frame, textvariable=self.svtelef, width=20, font=('ARIAL', 10))
        self.entrytelef.grid(column=1, row=4, padx=5, pady=5)
        self.datos_originales = []

        self.svBuscHC = tk.StringVar()
        self.entryBuscHC = tk.Entry(self.scrollable_frame, textvariable=self.svBuscHC, width=20, font=('ARIAL', 10))
        self.entryBuscHC.grid(column=1, row=5, padx=5, pady=8)

        self.svBuscDNI = tk.StringVar()
        self.entryBuscDNI = tk.Entry(self.scrollable_frame, textvariable=self.svBuscDNI, width=20, font=('ARIAL', 10))
        self.entryBuscDNI.grid(column=1, row=6, padx=5, pady=8)

        # validación para aceptar solo números
        self.entryedad.config(validate="key", validatecommand=(self.root.register(self.validate_number), '%P'))
        self.entryDNI.config(validate="key", validatecommand=(self.root.register(self.validate_number), '%P'))
        self.entrytelef.config(validate="key", validatecommand=(self.root.register(self.validate_number), '%P'))
        self.entryBuscDNI.config(validate="key", validatecommand=(self.root.register(self.validate_number), '%P'))
        
        # validación para aceptar solo letras
        self.entryNombre.config(validate="key", validatecommand=(self.root.register(self.validate_letter), '%P'))
        self.entryApellido.config(validate="key", validatecommand=(self.root.register(self.validate_letter), '%P'))

        self.btnGuardar = tk.Button(self.scrollable_frame, text='Guardar', width=9, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#158645', cursor='hand2', activebackground='#35BD6F',command=self.guardar_datos)
        self.btnGuardar.grid(column=3, row=5, padx=2, pady=5)

        self.btnModificar = tk.Button(self.scrollable_frame, text='Modificar', width=9, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#1658A2', cursor='hand2', activebackground='#3D69F0',command=self.modificar_datos)
        self.btnModificar.grid(column=4, row=5, padx=2, pady=5)

        self.btnEliminar = tk.Button(self.scrollable_frame, text='Eliminar', width=9, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#D32F2F', cursor='hand2', activebackground='#E72D40',command=self.eliminar_datos_seleccionados)
        self.btnEliminar.grid(column=5, row=5, padx=2, pady=5)

        self.btnImportar = tk.Button(self.scrollable_frame, text='Importar', width=10, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#CF811E', cursor='hand2', activebackground='#E72D40',command=self.importar_desde_excel)
        self.btnImportar.grid(column=4, row=1, padx=5, pady=5)

        self.btnExportar = tk.Button(self.scrollable_frame, text='Exportar', width=10, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#CF811E', cursor='hand2', activebackground='#E72D40',command=self.exportar_a_excel)
        self.btnExportar.grid(column=4, row=3, padx=15, pady=5)

        self.btnBuscarhc = tk.Button(self.scrollable_frame, text='Buscar', width=10, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#0F1010', cursor='hand2', activebackground='#FFFEFE', command=self.buscar_por_hc)
        self.btnBuscarhc.grid(column=2, row=5, padx=15, pady=8)

        self.btnBuscardni = tk.Button(self.scrollable_frame, text='Buscar', width=10, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#0F1010', cursor='hand2', activebackground='#FFFEFE', command=self.buscar_por_dni)
        self.btnBuscardni.grid(column=2, row=6, padx=5, pady=8)

        self.btnLimpiarBuscar = tk.Button(self.scrollable_frame, text='Limpiar Busqueda', width=20, font=('Arial', 10, 'bold'),fg='#FFFEFE', bg='#6D00FD', cursor='hand2', activebackground='#FFFEFE', command=self.verificar_y_cargar_tabla)
        self.btnLimpiarBuscar.grid(column=4, row=6, padx=5, pady=8)       

    def crear_tabla(self):
        # Crear la tabla con el widget Treeview
        self.tabla = ttk.Treeview(self.frame_tabla, columns=("HC", "Fecha", "Edad", "Sexo", "Nombre", "Apellido", "FechaN", "Telefono", "DNI"))

        # Definir encabezados
        self.tabla.heading("#0", text="ID")
        self.tabla.heading("HC", text="Historia Clínica")
        self.tabla.heading("Fecha", text="Fecha")
        self.tabla.heading("Edad", text="Edad")
        self.tabla.heading("Sexo", text="Sexo")
        self.tabla.heading("Nombre", text="Nombre")
        self.tabla.heading("Apellido", text="Apellido")
        self.tabla.heading("FechaN", text="Fecha de Nacimiento")
        self.tabla.heading("Telefono", text="Teléfono")
        self.tabla.heading("DNI", text="DNI")

        # Configurar las columnas
        self.tabla.column("#0", stretch=tk.NO, width=0)  # ID
        self.tabla.column("HC", anchor=tk.CENTER, width=100)
        self.tabla.column("Fecha", anchor=tk.CENTER, width=100)
        self.tabla.column("Edad", anchor=tk.CENTER, width=50)
        self.tabla.column("Sexo", anchor=tk.CENTER, width=50)
        self.tabla.column("Nombre", anchor=tk.W, width=100)
        self.tabla.column("Apellido", anchor=tk.W, width=100)
        self.tabla.column("FechaN", anchor=tk.CENTER, width=100)
        self.tabla.column("Telefono", anchor=tk.CENTER, width=80)
        self.tabla.column("DNI", anchor=tk.CENTER, width=80)

        # Configurar la barra de desplazamiento
        scrollbar_y = ttk.Scrollbar(self.frame_tabla, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscroll=scrollbar_y.set)
        scrollbar_y.grid(row=0, column=1, sticky="ns")

        # Ubicar la tabla en el grid
        self.tabla.grid(row=0, column=0, sticky="nsew")

        # Configurar la barra de desplazamiento horizontal
        self.frame_tabla.columnconfigure(0, weight=1)
        self.frame_tabla.rowconfigure(0, weight=1)

        # Manejar la selección de fila
        self.tabla.bind("<ButtonRelease-1>", self.on_select)
        self.cargar_formulario()

    def actualizar_tabla(self):
        # Limpiar la tabla antes de actualizar
        self.tabla.delete(*self.tabla.get_children())

        # Iterar sobre los datos originales y agregarlos a la tabla
        for paciente in self.datos_originales:
            row_data = [paciente[field] for field in ['HC', 'Fecha', 'Edad', 'Sexo', 'Nombre', 'Apellido', 'FechaN', 'Telefono', 'DNI']]
            self.tabla.insert("", "end", values=row_data, tags=(paciente['HC'],))

    def on_select(self, event):
        # Obtén el elemento seleccionado
        selected_item = self.tabla.selection()

        if selected_item:
            # Obtiene el valor de la Historia Clínica (HC) de la fila seleccionada
            hc_value = self.tabla.item(selected_item, 'values')[0]

            # Hacer algo con la Historia Clínica (HC), como imprimirlo
            print(f"Selected HC: {hc_value}")

            # También puedes acceder a los valores de la fila directamente desde el árbol
            values = self.tabla.item(selected_item, 'values')
            print(f"Values: {values}")

    def guardar_datos(self):
        try:
            hc = self.svHC.get()
            fecha = self.svfecha.get()
            edad = self.svedad.get()
            sexo = 'Masculino' if self.svsexo_m.get() else 'Femenino'
            nombre = self.svNombre.get()
            apellido = self.svApellido.get()
            fecha_nacimiento = self.svFechaN.get()
            telefono = self.svtelef.get()
            dni = self.svDNI.get()
        
            # Validar campos antes de guardar
            if not hc or not fecha or not edad or not sexo or not nombre or not apellido or not fecha_nacimiento or not telefono or not dni:
                messagebox.showwarning("Advertencia", "Todos los campos son obligatorios.")
                return

            nuevo_paciente = {
                'HC': hc,
                'Fecha': fecha,
                'Edad': edad,
                'Sexo': sexo,
                'Nombre': nombre,
                'Apellido': apellido,
                'FechaN': fecha_nacimiento,
                'Telefono': telefono,
                'DNI': dni
            }

            # Buscar si ya existe un paciente con el mismo HC y fecha
            paciente_existente = None
            for idx, paciente in enumerate(self.datos_originales):
                if paciente['HC'] == hc and paciente['Fecha'] == fecha:
                    paciente_existente = idx
                    break

            # Si se encuentra un paciente existente, reemplazar sus datos
            if paciente_existente is not None:
                self.datos_originales[paciente_existente] = nuevo_paciente
            else:
                # Si no se encuentra, agregar el nuevo paciente a la lista
                self.datos_originales.append(nuevo_paciente)

            # Insertar en la tabla y guardar en el archivo CSV
            row_data = [nuevo_paciente[field] for field in ['HC', 'Fecha', 'Edad', 'Sexo', 'Nombre', 'Apellido', 'FechaN', 'Telefono', 'DNI']]
            self.tabla.insert("", "end", values=row_data, tags=(hc,))
            self.guardar_en_csv(row_data)  # Aquí pasas los datos que deseas guardar al método

            # Limpiar los campos después de guardar
            for entry in [self.entryHC, self.entryfecha, self.entryedad, self.entryNombre, self.entryApellido, self.entryFechaN, self.entrytelef, self.entryDNI]:
                entry.delete(0, tk.END)

            self.checkbox_masculino.deselect()
            self.checkbox_femenino.deselect()

        except Exception as e:
            print("Error", f"Error al guardar datos: {str(e)}")
        self.actualizar_tabla()

    def guardar_en_csv(self, datos):
        try:
            with open("datos_adultos.csv", mode="w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["Historia Clínica", "Fecha", "Edad", "Sexo", "Nombres", "Apellidos", "Fecha de Nacimiento", "Teléfono", "DNI"])
                for paciente in self.datos_originales:
                    writer.writerow([paciente[field] for field in ['HC', 'Fecha', 'Edad', 'Sexo', 'Nombre', 'Apellido', 'FechaN', 'Telefono', 'DNI']])
        except Exception as e:
            print("Error", f"Error al guardar en CSV: {str(e)}")

    def cargar_formulario(self):
        try:
            with open("datos_adultos.csv", mode="r", encoding="utf-8") as file:
                reader = csv.reader(file)
                next(reader, None)
                datos = [row for row in reader]
                print("Datos leídos desde el archivo CSV:", datos)  # Imprime los datos para depurar

                # Limpiar la tabla antes de cargar los datos ordenados
                self.tabla.delete(*self.tabla.get_children())

                for row in datos:
                    self.tabla.insert("", "end", values=row, tags=(row[0],))
                    self.datos_originales.append({  # Asegúrate de agregar los datos al atributo datos_originales
                        'HC': row[0],
                        'Fecha': row[1],
                        'Edad': row[2],
                        'Sexo': row[3],
                        'Nombre': row[4],
                        'Apellido': row[5],
                        'FechaN': row[6],
                        'Telefono': row[7],
                        'DNI': row[8] 
                    })

        except FileNotFoundError:
            print("Archivo no encontrado")  # Agregado para depuración

    def eliminar_datos_seleccionados(self):
        selected_item = self.tabla.selection()

        if selected_item:
            hc_value = self.tabla.item(selected_item, 'values')[0]

            # Eliminar de datos_originales
            self.datos_originales = [paciente for paciente in self.datos_originales if paciente['HC'] != hc_value]

            # Eliminar de la tabla
            self.tabla.delete(selected_item)

            # Actualizar archivo CSV
            self.actualizar_archivo_csv()

    def buscar_por_hc(self):
        hc_buscar = self.svBuscHC.get()

        if not hc_buscar:
            self.cargar_tabla()
            messagebox.showwarning("Advertencia", "Por favor, ingrese la Historia Clínica para buscar.")
            return

        print("Historia Clínica a buscar:", hc_buscar)

        # Limpiar la tabla antes de mostrar el resultado
        self.tabla.delete(*self.tabla.get_children())

        # Buscar en datos_originales
        print("Datos originales:", self.datos_originales)
        pacientes_encontrados = [paciente for paciente in self.datos_originales if paciente['HC'] == hc_buscar]

        print("Pacientes encontrados:", pacientes_encontrados)

        if pacientes_encontrados:
            # Insertar los pacientes encontrados en la tabla
            for paciente_encontrado in pacientes_encontrados:
                row_data = [paciente_encontrado[field] for field in ['HC', 'Fecha', 'Edad', 'Sexo', 'Nombre', 'Apellido', 'FechaN', 'Telefono', 'DNI']]
                self.tabla.insert("", "end", values=row_data, tags=(hc_buscar,))
        else:
            messagebox.showinfo("Información", "No se encontraron resultados para la Historia Clínica proporcionada.")
        
        # Verificar si entryBuscHC o svBuscHC está vacío
        if not self.svBuscHC.get():
            # Si está vacío, cargar la tabla con todos los datos originales
            self.cargar_tabla()

    def buscar_por_dni(self):
        dni_buscar = self.svBuscDNI.get()

        if not dni_buscar:
            messagebox.showwarning("Advertencia", "Por favor, ingrese el DNI para buscar.")
            return

        # Limpiar la tabla antes de mostrar el resultado
        self.tabla.delete(*self.tabla.get_children())

        # Buscar en datos_originales
        pacientes_encontrados = [paciente for paciente in self.datos_originales if paciente['DNI'] == dni_buscar]

        if pacientes_encontrados:
            # Insertar los pacientes encontrados en la tabla
            for paciente_encontrado in pacientes_encontrados:
                row_data = [paciente_encontrado[field] for field in ['HC', 'Fecha', 'Edad', 'Sexo', 'Nombre', 'Apellido', 'FechaN', 'Telefono', 'DNI']]
                self.tabla.insert("", "end", values=row_data, tags=(dni_buscar,))
        else:
            messagebox.showinfo("Información", "No se encontraron resultados para el DNI proporcionado.")
  
    def filtrar_tabla(self):
        hc_buscar = self.svBuscHC.get()
        dni_buscar = self.svBuscDNI.get()

        self.tabla.delete(*self.tabla.get_children())

        for datos_fila in self.datos_originales:
            hc_fila = datos_fila[0]
            dni_fila = datos_fila[8]  # Ajustar el índice según tu estructura de datos

            if ((not hc_buscar or hc_fila == hc_buscar) and 
                (not dni_buscar or dni_fila == dni_buscar)):
                self.tabla.insert("", "end", values=datos_fila, tags=(datos_fila[0],))

    def cargar_tabla(self):
    # Limpiar la tabla antes de cargar todos los datos originales
        self.tabla.delete(*self.tabla.get_children())

        # Insertar todos los datos originales en la tabla
        for paciente in self.datos_originales:
            row_data = [paciente[field] for field in ['HC', 'Fecha', 'Edad', 'Sexo', 'Nombre', 'Apellido', 'FechaN', 'Telefono', 'DNI']]
            self.tabla.insert("", "end", values=row_data, tags=(paciente['HC'],))

    def verificar_y_cargar_tabla(self):
        # Verificar si svBuscHC o entryBuscHC está vacío
        if not self.svBuscHC.get():
            # Si está vacío, cargar la tabla con todos los datos originales
            self.cargar_tabla()
            return

        # Verificar si svBuscDNI o entryBuscDNI está vacío
        if not self.svBuscDNI.get():
            # Si está vacío, cargar la tabla con todos los datos originales
            self.cargar_tabla()
            return      

    def exportar_a_excel(self):
        try:
            # Crear un libro de Excel y una hoja de cálculo
            wb = openpyxl.Workbook()
            sheet = wb.active

            # Agregar encabezados
            encabezados = ["Historia Clínica", "Fecha", "Edad", "Sexo", "Nombres", "Apellidos", "Fecha de Nacimiento", "Teléfono", "DNI"]
            sheet.append(encabezados)

            # Agregar datos
            for paciente in self.datos_originales:
                fila_datos = [paciente[field] for field in ['HC', 'Fecha', 'Edad', 'Sexo', 'Nombre', 'Apellido', 'FechaN', 'Telefono', 'DNI']]
                sheet.append(fila_datos)

            # Guardar el archivo Excel
            archivo_excel = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")])
            if archivo_excel:
                wb.save(archivo_excel)
                messagebox.showinfo("Éxito", "Datos exportados correctamente a Excel.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel: {str(e)}")

    def importar_desde_excel(self):
        try:
            # Abrir el cuadro de diálogo para seleccionar el archivo Excel
            archivo_excel = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
            if not archivo_excel:
                return  # El usuario canceló la selección

            # Cargar el libro de Excel
            wb = openpyxl.load_workbook(archivo_excel)
            sheet = wb.active

            # Obtener los datos de la hoja de cálculo
            datos_nuevos = []
            for fila in sheet.iter_rows(min_row=2, values_only=True):
                datos_fila = [str(valor) if valor else "" for valor in fila]
                datos_nuevos.append(datos_fila)

            # Verificar duplicados y agregar datos únicos
            for nuevo_paciente in datos_nuevos:
                hc_nuevo = nuevo_paciente[0]
                if not any(paciente['HC'] == hc_nuevo for paciente in self.datos_originales):
                    # No es un duplicado, agregar a datos_originales y a la tabla
                    self.datos_originales.append({
                        'HC': hc_nuevo,
                        'Fecha': nuevo_paciente[1],
                        'Edad': nuevo_paciente[2],
                        'Sexo': nuevo_paciente[3],
                        'Nombre': nuevo_paciente[4],
                        'Apellido': nuevo_paciente[5],
                        'FechaN': nuevo_paciente[6],
                        'Telefono': nuevo_paciente[7],
                        'DNI': nuevo_paciente[8]
                    })
                    self.tabla.insert("", "end", values=nuevo_paciente, tags=(hc_nuevo,))

            messagebox.showinfo("Éxito", "Datos importados correctamente desde Excel.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al importar desde Excel: {str(e)}")
    
    def modificar_datos(self):
        # Obtener la fila seleccionada
        selected_item = self.tabla.selection()
        if not selected_item:
            messagebox.showerror("Error", "Seleccione un paciente para modificar.")
            return

        # Obtener los valores de la fila seleccionada
        row = self.tabla.item(selected_item, 'values')
        if not row:
            messagebox.showerror("Error", "No se encontraron datos para modificar.")
            return

        # Actualizar los campos con los valores de la fila seleccionada
        self.svHC.set(row[0])
        self.svfecha.set(row[1])
        self.svedad.set(row[2])
        if row[3].lower() == 'masculino':
            self.svsexo_m.set(1)
        else:
            self.svsexo_f.set(1)
        self.svNombre.set(row[4])
        self.svApellido.set(row[5])
        self.svFechaN.set(row[6])
        self.svtelef.set(row[7])
        self.svDNI.set(row[8])
        # Actualizar el archivo CSV
        self.actualizar_archivo_csv()
        # Modificar los valores necesarios en la fila seleccionada
        sexo_modificado = 'Masculino' if self.svsexo_m.get() else 'Femenino'
        nueva_fila = (
            self.svHC.get(), 
            self.svfecha.get(), 
            self.svedad.get(), 
            sexo_modificado, 
            self.svNombre.get(), 
            self.svApellido.get(), 
            self.svFechaN.get(), 
            self.svtelef.get(), 
            self.svDNI.get()
        )
        self.tabla.item(selected_item, values=nueva_fila)

    def actualizar_archivo_csv(self):
        try:
            with open("datos_adultos.csv", mode="w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["Historia Clínica", "Fecha", "Edad", "Sexo", "Nombres", "Apellidos", "Fecha de Nacimiento", "Teléfono", "DNI"])
                for fila_tabla in self.tabla.get_children():
                    datos_fila = self.tabla.item(fila_tabla)['values']
                    writer.writerow(datos_fila)
        except Exception as e:
            print("Error", f"Error al guardar en CSV: {str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    app = VentanaAdultos(root)
    app.mainloop()